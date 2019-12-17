# !/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time    : 2019/11/28 16:56
# @Author  : Jackadam
# @Email   :
# @File    : main.py
# @Software: PyCharm
from openpyxl import Workbook

wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime

ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import Workbook

from openpyxl.styles import Border, Side

from openpyxl.reader.excel import load_workbook


def set_border(ws, cell_range):
    print(cell_range)
    # rows = ws.range(cell_range)
    # for row in rows:
    #     row[0].style.borders.left.border_style = Border.BORDER_THIN
    #     row[-1].style.borders.right.border_style = Border.BORDER_THIN
    # for c in rows[0]:
    #     c.style.borders.top.border_style = Border.BORDER_THIN
    # for c in rows[-1]:
    #     c.style.borders.bottom.border_style = Border.BORDER_THIN


# usage example:
ws = load_workbook('sample.xlsx').get_active_sheet()
set_border(ws, "C3:H10")
